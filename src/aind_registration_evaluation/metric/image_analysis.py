import tifffile as tiff
import numpy as np
from sklearn.metrics import mutual_info_score
from sklearn.metrics import normalized_mutual_info_score
import statsmodels.api as sm
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from aind_registration_evaluation.metric.cell_masks import CellMasks
import matplotlib.pyplot as plt
import io
from openpyxl.drawing.image import Image
from datetime import datetime


class ImageAnalysis():
    
    def __init__(self, 
                 image1, 
                 image2,
                 image1_mask,
                 image2_mask,
                 transformation_matrix=None,
                 maxCentroidDistance=10,
                 cell_match_count=True, 
                 full_image=True, 
                 roi_patches=True, 
                 full_image_matching_mask=True, 
                 roi_patches_matching_mask=True,
                 num_matches_plot=True, 
                 show_image=True,
                 nmi=True, 
                 mi=True):
        '''
        Initializes the ImageAnalysis class with the provided images, masks, and optional parameters for analysis.
        - Initializes the images and masks, optionally applies a transformation matrix to align the second image and its mask with the first image.
        - Sets up parameters to control which aspects of the analysis will be performed, including cell matching count, full image analysis, ROI patches, and matching masks.
        - Optionally includes metrics like NMI and MI in the analysis.

        Parameters
        ------------------------
        image1: numpy.ndarray or str
            The first image to analyze. If a string is provided, it is interpreted as a file path to the image.

        image2: numpy.ndarray or str
            The second image to analyze. If a string is provided, it is interpreted as a file path to the image.

        image1_mask: numpy.ndarray or str
            The mask corresponding to the first image. If a string is provided, it is interpreted as a file path to the mask.

        image2_mask: numpy.ndarray or str
            The mask corresponding to the second image. If a string is provided, it is interpreted as a file path to the mask.

        transformation_matrix: numpy.ndarray, optional
            A 4x4 transformation matrix to align `image2` with `image1`. If provided, `image2` and `image2_mask` will be transformed accordingly.

        maxCentroidDistance: float, optional, default=10
            The maximum distance between centroids of cells in the two images to consider them as matching.

        cell_match_count: bool, optional, default=True
            If True, the analysis will include the count of matching cells.

        full_image: bool, optional, default=True
            If True, the analysis will include the full images.

        roi_patches: bool, optional, default=True
            If True, the analysis will include regions of interest (ROI) patches.

        full_image_matching_mask: bool, optional, default=True
            If True, the analysis will include the full image matching masks.

        roi_patches_matching_mask: bool, optional, default=True
            If True, the analysis will include ROI patches matching masks.

        num_matches_plot: bool, optional, default=True
            If True, the analysis will include the plot of matching cells by centroid distance

        show_image: bool, optional, default=True
            If True, the analysis will include an image of the aligned images

        nmi: bool, optional, default=True
            If True, the analysis will include normalized mutual information (NMI) as a metric.

        mi: bool, optional, default=True
            If True, the analysis will include mutual information (MI) as a metric.
        '''

        if isinstance(image1, str): 
            self.image1 = tiff.imread(image1)
            self.image1_path = image1
        else: 
            self.image1 = image1
            self.image1_path = None

        if isinstance(image2, str): 
            self.image2 = tiff.imread(image2)
            self.image2_path = image2
        else: 
            self.image2 = image2
            self.image2_path = None

        self.image1 = self.image1.astype(np.int16)
        self.image2 = self.image2.astype(np.int16)
        print('Shape:', self.image1.shape, self.image2.shape)

        min_shape = np.minimum(self.image1.shape, self.image2.shape)
    
        # Clip both arrays to the smaller size
        self.image1 = self.image1[:min_shape[0], :min_shape[1], :min_shape[2]]
        self.image2 = self.image2[:min_shape[0], :min_shape[1], :min_shape[2]]

        self.cell_masks = CellMasks(image1_mask, image2_mask, transformation_matrix, maxCentroidDistance)

        if transformation_matrix is not None:
            transformation_matrix = self.args['transform_matrix']
            inverse_matrix = np.linalg.inv(transformation_matrix)
            self.image2 = affine_transform(self.image2, inverse_matrix)
            self.image2_mask = affine_transform(self.image2_mask, inverse_matrix)

        self.cell_match_count = cell_match_count
        self.full_image = full_image
        self.roi_patches = roi_patches
        self.full_image_matching_mask = full_image_matching_mask
        self.roi_patches_matching_mask = roi_patches_matching_mask
        self.num_matches_plot = num_matches_plot
        self.show_image = show_image
        self.nmi = nmi
        self.mi = mi

        self.matching_cell_plot = None
        self.aligned_image = None
        

    def run_cell_match_count(self, maxCentroidDistance=None):
        '''
        Retrieves matching patches and cell counts based on the centroid distance.

        Parameters
        ------------------------
        maxCentroidDistance : float or None
            Maximum centroid distance to consider for matching cells. If None, uses the instance's maxCentroidDistance.

        Returns
        ------------------------
        result_dict : dict
            A dictionary containing the number of cells in each image, the number of matching cells, and the patches with their volumes.
        '''
        return self.cell_masks.get_matching(maxCentroidDistance)

    def create_cell_match_plot(self):
        """
        Creates and returns a plot showing the number of cell matches as a function of the centroid distance threshold.

        If the plot has already been created, it simply returns the existing plot.

        Returns
        -------
        plt : matplotlib.pyplot
            The plot object showing the number of matches vs. centroid distance.
        """
        if self.matching_cell_plot is None:
            distances, num_matching_cells, max_matches = self.cell_masks.matching_cells_by_distance_plot()
            
            y = [n/max_matches for n in num_matching_cells]
            plt.plot(distances, y, marker='o')
            plt.xlabel('Centroid Distance')
            plt.ylabel('Number of Matches ')
            plt.title('Number of Matches vs Centroid Distance')
            plt.grid(True)

            self.matching_cell_plot = plt

        return self.matching_cell_plot


    def create_aligned_image(self):
        """
        Creates and returns a plot of an aligned image, showing the middle slice of two images overlaid in RGB.

        The function overlays the middle slice of `image1` and `image2`, normalizing each slice to the range [0, 1].
        The first channel of the RGB image corresponds to the normalized `image1`, and the second channel corresponds to the normalized `image2`.

        If the plot has already been created, it simply returns the existing plot.

        Raises
        ------
        ValueError
            If `image1` and `image2` do not have the same shape.

        Returns
        -------
        plt : matplotlib.pyplot
            The plot object showing the aligned middle slice of the images in RGB.
        """
        if self.aligned_image is None:
            if self.image1.shape != self.image2.shape:
                raise ValueError("Images must have the same shape.")

            z_middle = self.image1.shape[0] // 2

            slice1 = self.image1[z_middle]
            slice2 = self.image2[z_middle]

            slice1_norm = (slice1 - np.min(slice1)) / (np.max(slice1) - np.min(slice1))
            slice2_norm = (slice2 - np.min(slice2)) / (np.max(slice2) - np.min(slice2))

            rgb_image = np.zeros((slice1.shape[0], slice1.shape[1], 3), dtype=np.float32)

            rgb_image[..., 0] = slice1_norm

            rgb_image[..., 1] = slice2_norm

            plt.imshow(rgb_image)
            plt.title(f'Slice {z_middle}')
            plt.axis('off')

            self.aligned_image = plt

        return self.aligned_image



    def run_full_image(self, img1=None, img2=None, normalize=True, applyMask=True):
        '''
        Computes mutual information (MI) and normalized mutual information (NMI) between two images.
        
        Parameters
        ------------------------
        img1 : numpy array or None
            The first image to compare. If None, the instance's image1 is used.
        
        img2 : numpy array or None
            The second image to compare. If None, the instance's image2 is used.
        
        normalize : bool, optional
            If True, the mutual information scores are normalized by comparing with randomized permutations. Default is True.
        
        applyMask : bool, optional
            If True, applies a mask to remove zero-values from both images before comparison. Default is True.

        Returns
        ------------------------
        resultDict : dict
            A dictionary containing the MI and NMI scores. If normalize is True, it also contains normalized MI and NMI scores and random permutation scores.
        '''
        img1_eval = img1 if img1 is not None else self.image1
        img2_eval = img2 if img2 is not None else self.image2

        if applyMask:
            mask1 = img1_eval != 0
            mask2 = img2_eval != 0
            combined_mask = mask1 & mask2
            img1_eval = img1_eval[combined_mask]
            img2_eval = img2_eval[combined_mask]

        flat_1 = img1_eval.flatten()
        flat_2 = img2_eval.flatten()
        
        if normalize:
            if self.mi: rand_scores_mi = []
            if self.nmi: rand_scores_nmi = []
            for i in range(5):
                rand_1 = np.random.permutation(flat_1)
                rand_2 = np.random.permutation(flat_2)
                
                
                if self.mi: rand_scores_mi.append(mutual_info_score(rand_1, rand_2))
                if self.nmi: rand_scores_nmi.append(normalized_mutual_info_score(rand_1, rand_2, average_method='geometric'))

        resultDict = {}

        if self.mi:
            mi_dict = {}
            res_mi = mutual_info_score(flat_1, flat_2)
            mi_dict['MI']= res_mi
            if normalize:
                avg_mi = np.mean(rand_scores_mi)
                norm_mi = (res_mi - avg_mi)/(1 - avg_mi)
                mi_dict['RandomMI'] = rand_scores_mi
                mi_dict['NormalizedMI'] = norm_mi
            resultDict.update(mi_dict)

        if self.nmi:
            nmi_dict = {}
            res_nmi = normalized_mutual_info_score(flat_1, flat_2, average_method='geometric')
            nmi_dict['NMI']= res_nmi
            if normalize:
                avg_nmi = np.mean(rand_scores_nmi)
                norm_nmi = (res_nmi - avg_nmi)/(1 - avg_nmi)
                nmi_dict['RandomNMI'] = rand_scores_nmi
                nmi_dict['NormalizedNMI'] = norm_nmi
            resultDict.update(nmi_dict)

        return resultDict

    def run_patches(self, img1=None, img2=None, maxCentroidDistance=None, normalize=True, applyMask=False):
        '''
        Computes mutual information (MI) and normalized mutual information (NMI) for patches extracted from two images.

        Parameters
        ------------------------
        img1 : numpy array or None
            The first image to analyze. If None, the instance's image1 is used.
        
        img2 : numpy array or None
            The second image to analyze. If None, the instance's image2 is used.
        
        maxCentroidDistance : float or None
            Maximum centroid distance to consider for matching cells. If None, uses the instance's maxCentroidDistance.
        
        normalize : bool, optional
            If True, the mutual information scores are normalized by comparing with randomized permutations. Default is True.
        
        applyMask : bool, optional
            If True, applies a mask to remove zero-values from both images before comparison. Default is False.

        Returns
        ------------------------
        result_dict : dict
            A dictionary containing the weighted average and standard deviation of MI and NMI scores. If normalize is True, it also contains normalized MI and NMI scores.
        '''
        patch_dict = self.cell_masks.get_matching(maxCentroidDistance)
        image1 = self.image1 if img1 is None else img1
        image2 = self.image2 if img2 is None else img2

        if self.mi: patch_results_mi = []
        if self.nmi: patch_results_nmi = []
        if normalize:
            if self.mi: patch_results_mi_normalized = []
            if self.nmi: patch_results_nmi_normalized = []
        patch_weights = []
        for coords, vol in patch_dict['Patches'].items():
            patch1, patch2 = image1[coords[0]:coords[3],coords[1]:coords[4],coords[2]:coords[5]], image2[coords[0]:coords[3],coords[1]:coords[4],coords[2]:coords[5]]
            scores = self.run_full_image(img1=patch1, img2=patch2, normalize=normalize, applyMask=applyMask)

            patch_weights.append(vol)
            if self.mi: patch_results_mi.append(scores['MI'])
            if self.nmi: patch_results_nmi.append(scores['NMI'])

            if normalize:
                if self.mi: patch_results_mi_normalized.append(scores['NormalizedMI'])
                if self.nmi: patch_results_nmi_normalized.append(scores['NormalizedNMI'])
                

        result_dict = {}

        if self.mi:
            weighted_stats = sm.stats.DescrStatsW(patch_results_mi, weights=patch_weights)
            mi_dict = {
                'WeightedAvgMI': weighted_stats.mean,
                'WeightedStdMI': weighted_stats.std
            }
            if normalize:
                weighted_stats = sm.stats.DescrStatsW(patch_results_mi_normalized, weights=patch_weights)
                mi_dict['WeightedAvgMINormalized']= weighted_stats.mean 
                mi_dict['WeightedStdMINormalized']= weighted_stats.std 
            result_dict.update(mi_dict)
        if self.nmi:
            weighted_stats = sm.stats.DescrStatsW(patch_results_nmi, weights=patch_weights)
            nmi_dict = {
                'WeightedAvgNMI': weighted_stats.mean,
                'WeightedStdNMI': weighted_stats.std
            }
            if normalize:
                weighted_stats = sm.stats.DescrStatsW(patch_results_nmi_normalized, weights=patch_weights)
                nmi_dict['WeightedAvgNMINormalized']= weighted_stats.mean 
                nmi_dict['WeightedStdNMINormalized']= weighted_stats.std 
            result_dict.update(nmi_dict)
        
        return result_dict
    

    def run_roi_patches(self, maxCentroidDistance=None, normalize=True):
        '''
        Runs analysis on ROI patches, returning results with or without normalization.

        Parameters
        ------------------------
        maxCentroidDistance : float or None
            Maximum centroid distance to consider for matching cells. If None, uses the instance's maxCentroidDistance.
        
        normalize : bool
            Whether to normalize the mutual information (MI) and normalized mutual information (NMI) scores.

        Returns
        ------------------------
        result_dict : dict
            A dictionary containing the weighted average and standard deviation of MI and NMI scores for the ROI patches.
        '''
        return self.run_patches(maxCentroidDistance=maxCentroidDistance, normalize=False, applyMask=False)

    
    def run_full_image_matching_mask(self, maxCentroidDistance=None):
        '''
        Evaluates mutual information (MI) and normalized mutual information (NMI) between the full images of the matching masks.

        Parameters
        ------------------------
        maxCentroidDistance : float or None
            Maximum centroid distance to consider for matching cells. If None, uses the instance's maxCentroidDistance.

        Returns
        ------------------------
        result_dict : dict
            A dictionary containing MI and NMI scores between the matching masks of the full images.
        '''
        image1_matching_mask, image2_matching_mask = self.cell_masks.get_matching_mask_images(maxCentroidDistance)
        result_dict = {}
        flat_1, flat_2 = image1_matching_mask.flatten(), image2_matching_mask.flatten()

        if self.mi: 
            res_mi = mutual_info_score(flat_1, flat_2)
            result_dict['MI']= res_mi
        if self.nmi:
            res_nmi = normalized_mutual_info_score(flat_1, flat_2, average_method='geometric')
            result_dict['NMI']= res_nmi
        
        return result_dict


    def run_roi_patches_matching_mask(self, maxCentroidDistance=None):
        '''
        Evaluates patches using the matching masks of the images, without normalization and without applying additional masking.

        Parameters
        ------------------------
        maxCentroidDistance : float or None
            Maximum centroid distance to consider for matching cells. If None, uses the instance's maxCentroidDistance.

        Returns
        ------------------------
        dict
            Results from evaluating the ROI patches based on the matching masks. Includes MI and NMI metrics without normalization.
        '''
        matching_mask1, matching_mask2 = self.cell_masks.get_matching_mask_images(maxCentroidDistance)
        return self.run_patches(img1=matching_mask1, img2=matching_mask2, maxCentroidDistance=maxCentroidDistance, normalize=False, applyMask=False)

    def run(self, maxCentroidDistance=None):
        '''
        Executes the analysis based on the specified configurations and prints the results.

        Parameters
        ------------------------
        maxCentroidDistance : float or None
            Maximum centroid distance to consider for matching cells. If None, uses the instance's maxCentroidDistance.

        This method performs the following analyses:
        - Cell Match Count: Prints the count of cell matches based on the maximum centroid distance.
        - Full Image: Prints the mutual information (MI) and normalized mutual information (NMI) for the full images.
        - ROI Patches: Prints MI and NMI metrics for the ROI patches.
        - Full Image on Matching Mask: Prints MI and NMI metrics for full images using the matching masks.
        - ROI Patches on Matching Mask: Prints MI and NMI metrics for ROI patches using the matching masks.
        '''
        results = {}
        if self.cell_match_count:
            print('\n')
            print("Cell Match count")
            res = self.run_cell_match_count(maxCentroidDistance=maxCentroidDistance)
            print(res)
            results['Cell Match Count'] = res
            print('\n')
        if self.full_image:
            print('\n')
            print("Full Image")
            res = self.run_full_image()
            print(res)
            results['Full Image'] = res
            print('\n')
        if self.roi_patches:
            print('\n')
            print("ROI Patches:")
            res = self.run_roi_patches(maxCentroidDistance=maxCentroidDistance)
            print(res)
            results['ROI Patches'] = res
            print('\n')
        if self.full_image_matching_mask:
            print('\n')
            print("Full Image on Matching Mask:")
            res = self.run_full_image_matching_mask(maxCentroidDistance=maxCentroidDistance)
            print(res)
            results["Full Image Matching Mask"] = res
            print('\n')
        if self.roi_patches_matching_mask:
            print('\n')
            print("ROI Patches on Matching Mask:")
            res = self.run_roi_patches_matching_mask(maxCentroidDistance=maxCentroidDistance)
            print(res)
            results["ROI Patches Matching Mask"] = res
            print('\n')

        return results
    
    def run_to_excel(self, maxCentroidDistance=None, file_path='/results/results.xlsx'):
        '''
        Runs the analysis and writes the results to an Excel file.

        Parameters
        ------------------------
        maxCentroidDistance : float or None
            Maximum centroid distance to consider for matching cells. If None, uses the instance's maxCentroidDistance.
        file_path : str
            Path to the Excel file where results will be saved. If the file already exists, new results will be appended after two empty rows.
        '''
        # Dictionary to store results
        results = {}

        if self.full_image:
            print("Start Full image")
            full_image_results = self.run_full_image()
            results["Full Image"] = {
                "MI": full_image_results.get('MI', None),
                "Normalized MI": full_image_results.get('NormalizedMI', None),
                "NMI": full_image_results.get('NMI', None),
                "Normalized NMI": full_image_results.get('NormalizedNMI', None)
            }
            
        if self.roi_patches:
            print("Start ROI")
            roi_patches_results = self.run_roi_patches(maxCentroidDistance=maxCentroidDistance)
            results["ROI Patches"] = {
                "MI": roi_patches_results.get('WeightedAvgMI', None),
                "NMI": roi_patches_results.get('WeightedAvgNMI', None),
            }
            
        if self.full_image_matching_mask:
            print("Start Full image on matching mask")
            full_image_mask_results = self.run_full_image_matching_mask(maxCentroidDistance=maxCentroidDistance)
            results["Full Image on Matching Mask"] = {
                "MI": full_image_mask_results.get('MI', None),
                "NMI": full_image_mask_results.get('NMI', None),
            }
            
        if self.roi_patches_matching_mask:
            print("Start ROI on matching mask")
            roi_patches_mask_results = self.run_roi_patches_matching_mask(maxCentroidDistance=maxCentroidDistance)
            results["ROI Patches on Matching Mask"] = {
                "MI": roi_patches_mask_results.get('WeightedAvgMI', None),
                "NMI": roi_patches_mask_results.get('WeightedAvgNMI', None),
            }

        # Create a DataFrame from the results dictionary
        df_results = pd.DataFrame(results).T

        if self.image1_path and self.image2_path:
            filenames = {
                " ": [""] * len(df_results),
                "  ": [""] * len(df_results),
                "Image1 path": [self.image1_path] + [""] * (len(df_results)-1),
                "Image2 path": [self.image2_path] + [""] * (len(df_results)-1),
            }
            df_filenames = pd.DataFrame(filenames, index=df_results.index)
            df_results = pd.concat([df_results, df_filenames], axis=1)


        # Create DataFrame for additional metrics
        if self.cell_match_count:
            print("Start Cell match count")
            cell_match_count_results = self.run_cell_match_count(maxCentroidDistance=maxCentroidDistance)
            additional_metrics = {
                " ": [""] * len(df_results),
                "  ": [""] * len(df_results),
                "Image1 # Cells": [cell_match_count_results.get('NumImg1Cells', None)] + [""] * (len(df_results)-1),
                "Image2 # Cells": [cell_match_count_results.get('NumImg2Cells', None)] + [""] * (len(df_results)-1),
                "Max Centroid Distance": [cell_match_count_results.get('maxCentroidDistance', None)] + [""] * (len(df_results)-1),
                "# Matching Cells": [cell_match_count_results.get('NumMatchingCells', None)] + [""] * (len(df_results)-1),
            }
            df_additional_metrics = pd.DataFrame(additional_metrics, index=df_results.index)
            df_results = pd.concat([df_results, df_additional_metrics], axis=1)

        new_sheet_name = 'Plot_' + datetime.now().strftime('%Y%m%d_%H%M%S')
        if self.num_matches_plot:
            sheet_info = {
                " ": [""] * len(df_results),
                "  ": [""] * len(df_results),
                "Plot Sheet Name": [new_sheet_name] + [""] * (len(df_results)-1)
            }
            df_sheet_info = pd.DataFrame(sheet_info, index=df_results.index)
            df_results = pd.concat([df_results, df_sheet_info], axis=1)

        if os.path.exists(file_path):
            # Load the existing workbook
            book = load_workbook(file_path)
            
            # Get the 'Results' sheet
            if 'Results' in book.sheetnames:
                sheet = book['Results']
                start_row = sheet.max_row + 2
            else:
                sheet = book.create_sheet('Results')
                start_row = 1
        else:
            # Create a new workbook if the file doesn't exist
            book = Workbook()
            sheet = book.active
            sheet.title = 'Results'
            start_row = 1

        sheet.cell(row=start_row, column=1, value="Metric")  # Add row name header
        for c_idx, col_name in enumerate(df_results.columns, start=2):
            sheet.cell(row=start_row, column=c_idx, value=col_name)
        start_row += 1

        # Write data including row names
        for r_idx, (row_name, row) in enumerate(df_results.iterrows(), start=start_row):
            sheet.cell(row=r_idx, column=1, value=row_name)  # Write row name
            for c_idx, value in enumerate(row, start=2):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        if self.num_matches_plot:
            p = self.create_cell_match_plot()

            # Save the plot to a BytesIO object
            img_buffer = io.BytesIO()
            p.savefig(img_buffer, format='png')
            img_buffer.seek(0)
            p.close()

            # Create a new sheet for the plot
            plot_sheet = book.create_sheet(title=new_sheet_name)
            img = Image(img_buffer)
            img.anchor = 'A1'
            plot_sheet.add_image(img)

        if self.show_image:
            i = self.create_aligned_image()

            img_buffer = io.BytesIO()
            i.savefig(img_buffer, format='png')
            img_buffer.seek(0)
            i.close()

            # Create a new sheet for the plot
            if new_sheet_name in book.sheetnames: plot_sheet = book[new_sheet_name]
            else: plot_sheet = book.create_sheet(title=new_sheet_name)
            
            img = Image(img_buffer)
            img.anchor = 'L1'
            plot_sheet.add_image(img)
            

        book.save(file_path)



